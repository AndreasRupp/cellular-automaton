from __future__ import print_function
from datetime import datetime

import os, sys
import numpy as np
import itertools
import xlsxwriter
import math 
import multiprocessing
import openpyxl
import glob, os
def cam_particles( domain_size, n_steps, jump_parameter, distribution,porosity, type_g, charges_g,type_i, charges_i, filename = 'output/cam.png', debug_mode=False):
  start_time = datetime.now()
  #print("Starting time is", start_time)

  if not os.path.exists(folder + '/raw'):  os.makedirs(folder+ '/raw')
  workbook = xlsxwriter.Workbook(folder + '/raw/'+ filename + '.xlsx', {'nan_inf_to_errors': True})

  worksheet = workbook.add_worksheet()
  eval_measures_names = ['n_solids',
            'n_particles',
           'n_composites',
            'n_solids_in_composites',
            'n_solids_in_discrete_BUs',
            'mean_particle_size', 
            'specific_surface_area',
            'contact_area_per_volume',
            'mean_diameter']
  worksheet.write(0, 0, 'time_step')
  column = 1         
  for name in eval_measures_names:
    worksheet.write(0,column, name)
    column += 1          


  try:
    import CAM
  except (ImportError, ModuleNotFoundError) as error:
    sys.path.append(os.path.dirname(os.path.abspath(__file__)) + os.sep  + ".." + os.sep + "import")
    import CAM

  try:
    from plot import plot, plot_to_file, plot_to_vtk
  except (ImportError, ModuleNotFoundError) as error:
    sys.path.append(os.path.dirname(os.path.abspath(__file__)) + os.sep  + ".." + os.sep + 
      "python_functions")
    from plot import plot, plot_to_file, plot_to_vtk

  
  const                 = CAM.config()
  const.nx              = domain_size
  const.debug_mode      = debug_mode
  PyCAM = CAM.include(const)
  Domain = PyCAM(jump_parameter)
  # print(const.nx)
  n_fields = np.prod(const.nx)
  n_solids = round((1 - porosity) * n_fields)
  # print('number of fields',n_fields, 'number of solids', n_solids)



  placement_i = n_solids * distribution
  placement_g = n_solids * (1- distribution)
  #print('Number cells of goethite', placement_g,'Number cells of illite', placement_i)
  #print(type_g)
  #print(type_i)
  size_g = np.prod(type_g)
  size_i = np.prod(type_i)
  n_g = round(placement_g/size_g)
  n_i = round(placement_i/size_i)
  # print('size g', size_g, 'size_i', size_i)
  #print('number of goethite',n_g,'number of illite',n_i)

  def stencil_size(const_stencil, area):
    return np.floor((const_stencil/(np.sqrt(area))) + 0.5)

  def diameter(PS):
      return 2 * math.sqrt(PS * (0.025 * 1000) ** 2 / math.pi)


  #typ 1
  summe = 0
  success = 0
  jp_g = stencil_size(jump_parameter,size_g)
  while success < np.floor(n_g/2):
    success = success + Domain.place_plane(jp_g, type_g, -1, charges_g )#y #[1] * 4 
  #print(success)
  summe = summe + success

  success = 0
  type_g = list(np.roll(type_g,1))
  charges_g = list(np.roll(charges_g,2))
  while success <  np.ceil(n_g/2):
    success = success + Domain.place_plane(jp_g,type_g,  -1, charges_g )#x #[1] * 4 
  #print(success)
  summe = summe + success

  # placement_i = n_solids - summe * size_g
  # n_i = np.floor(placement_i/size_i)


  #type 2
  jp_i = stencil_size(jump_parameter,size_i)
  success = 0
  while success <  np.floor(n_i/2):
    success = success + Domain.place_plane(jp_i, type_i, -1,   charges_i)#y_neg #[-1] * 4
  #print(success)
  summe = summe + success

  success = 0
  type_i = list(np.roll(type_i,1))
  charges_i = list(np.roll(charges_i,2))
  while success < np.ceil(n_i/2):
    success = success + Domain.place_plane(jp_i,type_i ,-1,charges_i)#x_neg #[-1] * 4
  #print(success)
  summe = summe + success
  #print('porosity ', (n_g * size_g + n_i * size_i)/n_fields)
  #print('jump_parameter goethite', jp_g,'jump_parameter illite', jp_i)

  # Domain.print_array()

  # Domain.place_particles()

  # save_data = np.zeros( (n_steps + 1, np.prod(const.nx)) ) 
  plot_data = np.zeros( (2, np.prod(const.nx)) ) 

  # save_data[0] = Domain.fields()
  plot_data[0] = Domain.fields()

  eval_data = Domain.eval_measures()
  eval_data[len(eval_measures_names)-1] = diameter(eval_data[5])
  worksheet.write(1, 0, 0)
  for i in range(len(eval_measures_names)):
    worksheet.write(1, i + 1, eval_data[i])

  # simulation    
  #print("start simulation")
  for step in range(n_steps):
    Domain.do_cam()
    if(step < 500  or (step % 50 == 0) or step == (n_steps-1)):
      eval_data = Domain.eval_measures()
      eval_data[len(eval_measures_names)-1] = diameter(eval_data[5])
      worksheet.write(step + 2, 0, step+1)
      for i in range(len(eval_measures_names)):
        worksheet.write(step + 2, i+1, eval_data[i])
    # save_data[step+1] = Domain.fields()
    # print(step + 1)

  plot_data[1] = Domain.fields()
  

  #print('specific_surface_area',eval_data[7])
  #print('contact_area_per_volume', eval_data[8])
  # print(Domain.average_particle_size())
  # print(Domain.bulk_distance(save_data[0],save_data[-1]))
  end_time = datetime.now() 
  
  # save_data[(save_data <= n_g) & (save_data > 0)] = 1
  # save_data[save_data > n_g] = 2

  plot_data[(plot_data <= n_g) & (plot_data > 0)] = 1
  plot_data[plot_data > n_g] = 2
  
  # plot_to_vtk("output/cam", [save_data[0]], const.nx)
  # if not os.path.exists('output2D'):  os.makedirs('output2D')
  # plot_to_vtk("output/cam", save_data, const.nx)
  # if not os.path.exists('output1'):  os.makedirs('output1')
  # plot_to_vtk("output1/cam", [save_data[-1]], const.nx)
  #print("done")
  text = 'SSA ' + str(round(eval_data[6],2)) + ' (L^-1), CA/V ' + str(round(eval_data[7],2)) + ' (L^-1)' + ' mean diamter ' + str(round(diameter(eval_data[5]))) + ' nm'
  # print(filename)
  plot_to_file(const.nx, plot_data[-1], folder + filename + '.png', text)
  # plot(const.nx, plot_data, 0)
  workbook.close()
  print(filename)
  #print("Program ended at", end_time, "after", end_time-start_time)
  
class setting:
  def __init__(self,
    domain_size = [50,50],
    n_steps = 5000, 
    jump_parameter = 20, 
    distribution= 0.5,
    porosity = 0.9, 
    type_g = [2, 34], 
    charges_g = [1] * 4 ,
    type_i = [2, 6], 
    charges_i = [-1] * 4 , 
    filename = 'filename', 
    debug_mode= False):
    self.domain_size = domain_size
    self.n_steps = n_steps
    self.jump_parameter = jump_parameter
    self.distribution = distribution
    self.porosity = porosity
    self.type_g = type_g
    self.charges_g = charges_g
    self.type_i = type_i
    self.charges_i = charges_i
    self.filename = filename
    self.debug_mode = debug_mode 
def run_test_from_class(s):
  cam_particles( s.domain_size, s.n_steps, s.jump_parameter, s.distribution, s.porosity, s.type_g, s.charges_g, s.type_i, s.charges_i, s.filename, s.debug_mode)
# --------------------------------------------------------------------------------------------------
# Function main.
# --------------------------------------------------------------------------------------------------
def main(d):
  #d = 500
  debug_mode = False
  domain_size = [d,d]
  n_steps = 5000
  porosity = 0.9
  jump_const = 20
  type_i_addition = 0.95


  illite_fine = [2, 6]
  illite_medium = [2, 30]
  goethite_fine = [1, 17]
  goethite_coarse = [2, 34]

  uniform_positive = [1] * 4 
  uniform_negative = [-1] * 4 

  face_to_edge = [1,1, -1, -1]
  face_to_face = [-1,-1, -1, -1]
  #goethite positive charges
  #basal of illite = long edges
  #illite negative charges on basal planes and pos or negative charges on other egdes

  y = [0 ,0 , 1 ,1]
  y_neg = [0 ,0 , -1 ,-1]
  x =  [ 1,1,0,0]
  x_neg =  [ -1,-1,0,0]

  fun_args  = []

  type_i_addition = 1- 0.05
  filename = 'goethite_illite_5'
  fun_args.append(setting( domain_size, n_steps,jump_const,type_i_addition, porosity,  goethite_coarse,uniform_positive, illite_fine,uniform_negative ,filename,debug_mode))
  #cam_particles(domain_size, n_steps,jump_const,type_i_addition, porosity,  goethite_coarse,uniform_positive, illite_fine,uniform_negative ,filename,debug_mode)
  type_i_addition = 1 - 0.45
  filename = 'goethite_illite_45'
  fun_args.append(setting(domain_size, n_steps,jump_const,type_i_addition, porosity,  goethite_coarse,uniform_positive, illite_fine,uniform_negative ,filename,debug_mode))
  # cam_particles(domain_size, n_steps,jump_const,type_i_addition, porosity,  goethite_coarse,uniform_positive, illite_fine,uniform_negative ,filename,debug_mode)

  type_i_addition = 1 - 0.55
  filename = 'goethite_illite_55'
  fun_args.append(setting(domain_size, n_steps,jump_const,type_i_addition, porosity,  goethite_coarse,uniform_positive, illite_fine,uniform_negative ,filename,debug_mode))
  # cam_particles(domain_size, n_steps,jump_const,type_i_addition, porosity,  goethite_coarse,uniform_positive, illite_fine,uniform_negative ,filename,debug_mode)
  type_i_addition = 1 - 0.95
  filename = 'goethite_illite_95'
  fun_args.append(setting( domain_size, n_steps,jump_const,type_i_addition, porosity,  goethite_coarse,uniform_positive, illite_fine,uniform_negative ,filename,debug_mode))
  # cam_particles(domain_size, n_steps,jump_const,type_i_addition, porosity,  goethite_coarse,uniform_positive, illite_fine,uniform_negative ,filename,debug_mode)



  type_i_addition = 0.5
  filename = 'illite_medium_edge_to_face'
  fun_args.append(setting( domain_size, n_steps,jump_const,type_i_addition, porosity,  illite_medium, face_to_edge, illite_medium, face_to_edge ,filename, debug_mode))
  # cam_particles(domain_size, n_steps,jump_const,type_i_addition, porosity,  illite_medium, face_to_edge, illite_medium, face_to_edge ,filename, debug_mode)
  filename = 'illite_fine_edge_to_face'
  fun_args.append(setting(  domain_size, n_steps,jump_const,type_i_addition, porosity,  illite_fine, face_to_edge, illite_fine, face_to_edge ,filename, debug_mode))
  #cam_particles(domain_size, n_steps,jump_const,type_i_addition, porosity,  illite_fine, face_to_edge, illite_fine, face_to_edge ,filename, debug_mode)

  filename = 'illite_medium_face_to_face'
  fun_args.append(setting( domain_size, n_steps,jump_const,type_i_addition, porosity,  illite_medium, face_to_face, illite_medium, [i * (-1) for i in face_to_face],filename,debug_mode))
  # cam_particles(domain_size, n_steps,jump_const,type_i_addition, porosity,  illite_medium, face_to_face, illite_medium, [i * (-1) for i in face_to_face],filename,debug_mode)
  filename = 'illite_fine_face_to_face'
  fun_args.append(setting( domain_size, n_steps,jump_const,type_i_addition, porosity,  illite_fine, face_to_face, illite_fine, [i * (-1) for i in face_to_face] ,filename, debug_mode))
  #cam_particles(domain_size, n_steps,jump_const,type_i_addition, porosity,  illite_fine, face_to_face, illite_fine, [i * (-1) for i in face_to_face] ,filename, debug_mode)

  processes = []
  for fun_arg in fun_args:
    t = multiprocessing.Process(target=run_test_from_class, args=(fun_arg,))
    processes.append(t)
    t.start()
    # time.sleep(2)


  # while multiprocessing.active_children():
  #   val = input("Enter your value: ")
  #   if val == "kill_all_children":
  #     active = multiprocessing.active_children()
  #     for child in active:
  #       child.kill()
  #     time.sleep(2)
  print('fertisch?')     
  for one_process in processes:
    one_process.join()
    # time.sleep(2)

   
  print('fertisch')


# --------------------------------------------------------------------------------------------------
# Define main function.
# -------------------------------------------------------------------------------------------------- 
if __name__ == "__main__":
  domain_sizes = [500]#[100,250,500,750,1000,1250,1500,1750,2000,2250,2500]
  for size in domain_sizes:
    folder = 'output_2D_rotated_composites/' + str(size) + '/' 
    if not os.path.exists(folder):  os.makedirs(folder)
    main(size) #len(sys.argv) > 1 and sys.argv[1] == "True"


    # workbook = xlsxwriter.Workbook(folder + 'simuation_eval.xlsx')#,, { 'constant_memory': True}  
    
    # for file in glob.glob(folder + "/raw/*.xlsx"):

    #   print(file)
    #   sheet_name = file.replace('.xlsx','')
    #   sheet_name = sheet_name.replace(folder,'')
    #   sheet_name = sheet_name.replace('/raw/','')
    #   print(sheet_name)
    #   worksheet = workbook.add_worksheet(sheet_name)
    #   # Define variable to load the dataframe
    #   dataframe = openpyxl.load_workbook(file)
   
    #   # Define variable to read sheet
    #   dataframe1 = dataframe.active
   
    #   # Iterate the loop to read the cell 
    #   row_n = 0
    #   for row in range(0, dataframe1.max_row):
    #     col_n = 0
    #     for col in dataframe1.iter_cols(1, dataframe1.max_column):
    #       #print(col[row].value)
    #       #print(row_n)
    #       #print(col_n)
    #       worksheet.write(row_n, col_n, col[row].value)
    #       col_n = col_n +1
    #     row_n = row_n + 1
    #   # os.remove(file)
    # workbook.close() 

